import socket
import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IP = '127.0.0.1'
PUERTO = 65432

def guardar(datos):
    try:
        nombre = datos['artista']
        lista = datos['datos']
        
        print("\nRecibido: " + nombre)
        
        df = pd.DataFrame(lista)
        archivo = "Viaje_" + nombre.replace(" ", "_") + ".xlsx"
        
        ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), archivo)
        
        df.to_excel(ruta, index=False)
        
        wb = load_workbook(ruta)
        ws = wb.active
        
        f = Font(bold=True, color="FFFFFF", size=11)
        bg = PatternFill("solid", fgColor="4F81BD")
        ali = Alignment(horizontal="center", vertical="center")
        borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for i in range(1, ws.max_column + 1):
            c = ws.cell(row=1, column=i)
            c.font = f
            c.fill = bg
            c.alignment = ali
            c.border = borde

        idx = 1
        for col in ws.columns:
            letra = get_column_letter(idx)
            ancho_max = 0
            for celda in col:
                celda.border = borde
                celda.alignment = ali
                val = str(celda.value)
                if len(val) > ancho_max: ancho_max = len(val)
            
            final = ancho_max + 4
            if final > 50: final = 50
            ws.column_dimensions[letra].width = final
            idx += 1

        wb.save(ruta)
        print(" > Excel guardado.")
        
    except:
        print(" Error al guardar el archivo.")

if __name__ == "__main__":
    print("=========================================")
    print("       SERVIDOR TourMate ACTIVO")
    print("     Pulsa Ctrl + C para apagarlo.")
    print("=========================================")
    
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind((IP, PUERTO))
    s.listen()
    s.settimeout(0.2) 
    
    try:
        while True:
            try:
                c, addr = s.accept()
                c.settimeout(None)
                
                todo = b""
                while True:
                    trozo = c.recv(4096)
                    if not trozo: break
                    todo += trozo
                
                if todo:
                    txt = todo.decode('utf-8')
                    guardar(json.loads(txt))
                
                c.close()
                
            except socket.timeout:
                pass 
                
    except KeyboardInterrupt:
        print("\nApagado")
    
    s.close()